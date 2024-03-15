import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
from collections import Counter

def fetch_notes(url):
    # Fetch the webpage content with a custom user-agent header
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'}
    response = requests.get(url, headers=headers)

    # Check if the request was successful
    if response.status_code != 200:
        print(f"Error: Couldn't fetch the webpage. Status code: {response.status_code}")
        return None

    soup = BeautifulSoup(response.text, 'html.parser')

    # Find the div with class 'notes_list mb-2'
    notes_div = soup.find('div', class_='notes_list mb-2')

    # Check if notes_div is None
    if notes_div is None:
        print("Error: Couldn't find notes section on the webpage.")
        return None

    # Extract notes from the div and split by space
    all_notes = [note_span.text.strip() for note_span in notes_div.find_all('span', class_='nowrap')]
    return all_notes

def write_to_excel(perfume_name, all_notes):
    # Load existing workbook or create a new one if it doesn't exist
    filename = "collection_app.xlsx"
    try:
        wb = load_workbook(filename)
        ws = wb.active
        ws.append(["", "", ""])
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Perfume Collection"
        ws.append(["Perfume", "Notes", "Date Added"])
        ws.append(["", "", ""])

    # Get the current date
    date_added = datetime.now().strftime("%m/%d/%Y")

    # Check if the perfume already exists in the collection
    perfume_names = [row[0].value for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1)]
    if perfume_name in perfume_names:
        print(f"'{perfume_name}' already exists in the collection. Enter another link.")
        return

    # Add data to the workbook
    for note in all_notes:
        ws.append([perfume_name, f'"{note}"', date_added])
    #ws.append(["", "", ""])

    # Save the workbook
    wb.save(filename)
    print(f"Data for {perfume_name} has been written to {filename}")
    #wb.active.append(["", "", ""])

def find_top_notes(n):
    filename = "collection_app.xlsx"
    wb = load_workbook(filename)
    ws = wb.active
    notes_column = [cell.value.strip('"') for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2) for cell in row if cell.value]
    top_notes = Counter(notes_column).most_common(n)
    print(f"Top {n} most common notes in your collection:")
    for note, count in top_notes:
        print(f"{note}: {count} occurrences")

def main():
    while True:
        user_input = input("\nWelcome to the App!\nGo to parfumo.net and paste your Fragrance's URL down below to begin\nEnter a URL or a list of URLs separated by spaces (or 'done' to proceed): ").strip()
        if user_input.lower() == 'done':
            break
        
        urls = user_input.split()
        for url in urls:
            all_notes = fetch_notes(url)
            if all_notes:
                perfume_name = url.split('/')[-1]
                write_to_excel(perfume_name, all_notes)

    while True:
        command = input("Enter command (top 5 / top 10 / quit): ").strip().lower()
        filename = "collection_app.xlsx"
        if command == 'top 5':
            find_top_notes(5)
        elif command == 'top 10':
            find_top_notes(10)
        elif command == 'quit':
            os.system(filename)
            exit()
        else:
            print("Invalid command. Please enter 'top 5', 'top 10' or 'quit'.")

        # Ask for note name
        note_name = input("Type the note name to see which perfumes contain it: ").strip()
        if note_name:
            find_perfumes_with_note(note_name)

def find_perfumes_with_note(note_name):
    filename = "collection_app.xlsx"
    wb = load_workbook(filename)
    ws = wb.active
    perfume_names = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            perfume_name = cell.value
            if perfume_name:
                notes = cell.offset(column=1).value.strip('"')
                if note_name in notes:
                    perfume_names.append(perfume_name)
    if perfume_names:
        print(f"Perfumes containing the note '{note_name}':")
        for perfume in perfume_names:
            print(perfume)
    else:
        print(f"No perfumes found containing the note '{note_name}'.")

if __name__ == "__main__":
    main()

